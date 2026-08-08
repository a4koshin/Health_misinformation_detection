"""Batch dataset predictions via SomBERTb Task A (CSV / Excel)."""

from __future__ import annotations

import csv
import io
from typing import Any

TEXT_COLUMN_CANDIDATES = (
    "text",
    "input_text",
    "claim",
    "claim_text",
    "sentence",
    "content",
    "message",
)

MAX_DATASET_ROWS = 20_000
EMPTY_FILE_MESSAGE = "The uploaded file is empty. Add claim text and try again."
ALLOWED_EXTENSIONS = {
    ".csv",
    ".xlsx",
    ".xlsm",
    ".xls",
    ".xltx",
    ".xltm",
}


def _extension(filename: str) -> str:
    name = (filename or "").strip().lower()
    if "." not in name:
        return ""
    return "." + name.rsplit(".", 1)[-1]


def _pick_text_column(headers: list[str]) -> str | None:
    normalized = {header.strip().lower(): header for header in headers if header}
    for candidate in TEXT_COLUMN_CANDIDATES:
        if candidate in normalized:
            return normalized[candidate]
    if len(headers) == 1:
        return headers[0]
    return None


def _read_csv_rows(raw: bytes) -> list[str]:
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = [h or "" for h in (reader.fieldnames or [])]
    if not headers:
        # Headerless single-column file.
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        return lines
    column = _pick_text_column(headers)
    if not column:
        raise ValueError(
            "CSV must include a text column named text, input_text, claim, "
            "sentence, or content."
        )
    values: list[str] = []
    for row in reader:
        values.append(str(row.get(column) or "").strip())
    return values


def _read_excel_rows(raw: bytes, filename: str) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Excel support requires openpyxl. Install it with: pip install openpyxl"
        ) from exc

    ext = _extension(filename)
    if ext == ".xls":
        raise ValueError(
            "Legacy .xls is not supported. Save the file as .xlsx or .csv and try again."
        )

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        first = next(rows_iter, None)
        if first is None:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in first]
        # Headerless single-column sheet.
        if len(headers) == 1 and headers[0].strip().lower() not in TEXT_COLUMN_CANDIDATES:
            values = [headers[0]] if headers[0] else []
            for row in rows_iter:
                cell = row[0] if row else None
                if cell is None:
                    values.append("")
                else:
                    values.append(str(cell).strip())
            return values

        column = _pick_text_column(headers)
        if not column:
            raise ValueError(
                "Excel file must include a text column named text, input_text, "
                "claim, sentence, or content."
            )
        index = headers.index(column)
        values: list[str] = []
        for row in rows_iter:
            if not row or index >= len(row) or row[index] is None:
                values.append("")
            else:
                values.append(str(row[index]).strip())
        return values
    finally:
        workbook.close()


def extract_claim_texts(raw: bytes, filename: str) -> list[str]:
    ext = _extension(filename)
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError("Upload a .csv or Excel (.xlsx) file.")
    if ext == ".csv":
        return _read_csv_rows(raw)
    return _read_excel_rows(raw, filename)


def predict_dataset_file(raw: bytes, filename: str) -> dict[str, Any]:
    """Classify each claim with Task A. Does not call gatekeeper/LLM per row."""
    from services.predictor_service import predict_reliability_batch

    if not raw or not raw.strip():
        raise ValueError(EMPTY_FILE_MESSAGE)

    texts = extract_claim_texts(raw, filename)
    filled = [(index, text) for index, text in enumerate(texts, start=1) if text]
    if not filled:
        raise ValueError(EMPTY_FILE_MESSAGE)

    if len(texts) > MAX_DATASET_ROWS:
        raise ValueError(
            f"Dataset is too large ({len(texts)} rows). "
            f"Upload at most {MAX_DATASET_ROWS:,} rows."
        )

    predictions = predict_reliability_batch([text for _, text in filled])

    results: list[dict[str, Any]] = []
    reliable_count = 0
    misinformation_count = 0
    error_count = 0
    processed_rows = 0
    pred_by_row = {
        row_number: prediction
        for (row_number, _text), prediction in zip(filled, predictions)
    }

    for index, text in enumerate(texts, start=1):
        if not text:
            error_count += 1
            results.append(
                {
                    "row": index,
                    "text": "",
                    "prediction": None,
                    "error": "Empty row.",
                }
            )
            continue
        prediction = pred_by_row.get(index)
        if not prediction:
            error_count += 1
            results.append(
                {
                    "row": index,
                    "text": text,
                    "prediction": None,
                    "error": "Prediction failed.",
                }
            )
            continue
        label = prediction.get("label")
        if label == "Reliable":
            reliable_count += 1
        else:
            label = "Non-Reliable"
            misinformation_count += 1
        processed_rows += 1
        results.append(
            {
                "row": index,
                "text": text,
                "prediction": label,
                "error": None,
            }
        )

    return {
        "total_rows": len(texts),
        "processed_rows": processed_rows,
        "reliable_count": reliable_count,
        "misinformation_count": misinformation_count,
        "error_count": error_count,
        "results": results,
    }
