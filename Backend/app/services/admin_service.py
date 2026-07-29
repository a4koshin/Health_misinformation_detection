import csv
import io
import uuid
from pathlib import Path

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.security import hash_password
from app.models.detection import Detection
from app.models.user import User
from app.schemas.admin import (
    AdminUserCreate,
    AdminUserUpdate,
    DashboardStats,
    DatasetPredictionResponse,
    DatasetPredictionRow,
)
from app.services import auth_service, detection_service

TEXT_COLUMN_KEYS = ("text", "input_text", "claim", "sentence", "content")
CSV_EXTENSIONS = {".csv"}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}
SUPPORTED_EXTENSIONS = CSV_EXTENSIONS | EXCEL_EXTENSIONS


class UserNotFoundError(Exception):
    pass


class LastAdminError(Exception):
    pass


def list_users(db: Session) -> list[User]:
    return db.query(User).order_by(User.created_at.desc()).all()


def create_user(db: Session, data: AdminUserCreate) -> User:
    if auth_service.get_user_by_email(db, data.email):
        raise auth_service.EmailAlreadyRegisteredError()

    user = User(
        email=data.email,
        full_name=data.full_name,
        hashed_password=hash_password(data.password),
        role=data.role,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def update_user(
    db: Session,
    user_id: uuid.UUID,
    data: AdminUserUpdate,
    current_admin: User,
) -> User:
    user = auth_service.get_user_by_id(db, user_id)
    if user is None:
        raise UserNotFoundError()

    if data.email is not None and data.email != user.email:
        if auth_service.get_user_by_email(db, data.email):
            raise auth_service.EmailAlreadyRegisteredError()
        user.email = data.email

    if "full_name" in data.model_fields_set:
        user.full_name = data.full_name

    if data.password:
        user.hashed_password = hash_password(data.password)

    if data.role is not None and data.role != user.role:
        if user.role == "admin" and data.role != "admin":
            _ensure_another_admin_exists(db, user.id)
        user.role = data.role

    db.add(user)
    db.commit()
    db.refresh(user)

    if user.id == current_admin.id:
        db.refresh(current_admin)

    return user


def delete_user(
    db: Session,
    user_id: uuid.UUID,
    current_admin: User,
) -> None:
    user = auth_service.get_user_by_id(db, user_id)
    if user is None:
        raise UserNotFoundError()

    if user.id == current_admin.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="You cannot delete your own account.",
        )

    if user.role == "admin":
        _ensure_another_admin_exists(db, user.id)

    db.delete(user)
    db.commit()


def get_dashboard_stats(db: Session) -> DashboardStats:
    total_users = db.query(func.count(User.id)).scalar() or 0
    total_admins = (
        db.query(func.count(User.id)).filter(User.role == "admin").scalar() or 0
    )
    total_detections = db.query(func.count(Detection.id)).scalar() or 0
    reliable_count = (
        db.query(func.count(Detection.id))
        .filter(Detection.label == "Reliable")
        .scalar()
        or 0
    )
    misinformation_count = (
        db.query(func.count(Detection.id))
        .filter(Detection.label == "Misinformation")
        .scalar()
        or 0
    )
    pending_count = (
        db.query(func.count(Detection.id))
        .filter(Detection.label.is_(None))
        .scalar()
        or 0
    )

    return DashboardStats(
        total_users=total_users,
        total_admins=total_admins,
        total_detections=total_detections,
        reliable_count=reliable_count,
        misinformation_count=misinformation_count,
        pending_count=pending_count,
    )


def _find_text_header(headers: list[str]) -> str | None:
    normalized = {
        header.strip().lower(): header
        for header in headers
        if header and str(header).strip()
    }
    for key in TEXT_COLUMN_KEYS:
        if key in normalized:
            return normalized[key]
    return None


def _rows_from_csv(file_bytes: bytes) -> list[dict[str, object]]:
    try:
        decoded = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV file must be UTF-8 encoded.",
        ) from exc

    reader = csv.DictReader(io.StringIO(decoded))
    if not reader.fieldnames:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV file is empty or missing a header row.",
        )

    text_header = _find_text_header([str(header) for header in reader.fieldnames])
    if text_header is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must include a text column named text, input_text, claim, sentence, or content.",
        )

    return [
        {
            "row": index,
            "text": (row.get(text_header) or "").strip(),
        }
        for index, row in enumerate(reader, start=2)
    ]


def _rows_from_excel(file_bytes: bytes, extension: str) -> list[dict[str, object]]:
    if extension == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel .xls support is unavailable. Save the file as .xlsx or CSV.",
            ) from exc

        try:
            book = xlrd.open_workbook(file_contents=file_bytes)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Unable to read the .xls file. Try saving it as .xlsx or CSV.",
            ) from exc

        sheet = book.sheet_by_index(0)
        if sheet.nrows == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file is empty.",
            )

        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in sheet.row(0)
        ]
        text_header = _find_text_header(headers)
        if text_header is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File must include a text column named text, input_text, claim, sentence, or content.",
            )
        text_index = headers.index(text_header)

        rows: list[dict[str, object]] = []
        for index in range(1, sheet.nrows):
            value = sheet.cell_value(index, text_index)
            text = str(value).strip() if value is not None else ""
            rows.append({"row": index + 1, "text": text})
        return rows

    try:
        workbook = load_workbook(
            io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unable to read the Excel file. Use .xlsx, .xlsm, or CSV.",
        ) from exc

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file is empty.",
        ) from exc

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    text_header = _find_text_header(headers)
    if text_header is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must include a text column named text, input_text, claim, sentence, or content.",
        )
    text_index = headers.index(text_header)

    rows = []
    for index, row in enumerate(rows_iter, start=2):
        value = row[text_index] if text_index < len(row) else None
        text = str(value).strip() if value is not None else ""
        rows.append({"row": index, "text": text})
    return rows


def _load_dataset_rows(file_bytes: bytes, filename: str) -> list[dict[str, object]]:
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only CSV and Excel files are supported (.csv, .xlsx, .xlsm, .xls, .xltx, .xltm).",
        )

    if extension in CSV_EXTENSIONS:
        return _rows_from_csv(file_bytes)
    return _rows_from_excel(file_bytes, extension)


def predict_dataset(file_bytes: bytes, filename: str) -> DatasetPredictionResponse:
    rows = _load_dataset_rows(file_bytes, filename)

    results: list[DatasetPredictionRow] = []
    reliable_count = 0
    misinformation_count = 0
    error_count = 0

    for item in rows:
        raw_text = str(item["text"])
        row_number = int(item["row"])

        if not raw_text:
            error_count += 1
            results.append(
                DatasetPredictionRow(
                    row=row_number,
                    text="",
                    error="Empty text cell.",
                )
            )
            continue

        try:
            prediction = detection_service.predict(raw_text)
            if prediction == "Reliable":
                reliable_count += 1
            elif prediction in {"Misinformation", "Non-Reliable"}:
                misinformation_count += 1
            results.append(
                DatasetPredictionRow(
                    row=row_number,
                    text=raw_text,
                    prediction=prediction,
                )
            )
        except HTTPException as exc:
            error_count += 1
            results.append(
                DatasetPredictionRow(
                    row=row_number,
                    text=raw_text,
                    error=str(exc.detail),
                )
            )

    return DatasetPredictionResponse(
        total_rows=len(results),
        processed_rows=len(results) - error_count,
        reliable_count=reliable_count,
        misinformation_count=misinformation_count,
        error_count=error_count,
        results=results,
    )


def _ensure_another_admin_exists(db: Session, user_id: uuid.UUID) -> None:
    remaining_admins = (
        db.query(func.count(User.id))
        .filter(User.role == "admin", User.id != user_id)
        .scalar()
        or 0
    )
    if remaining_admins == 0:
        raise LastAdminError()
